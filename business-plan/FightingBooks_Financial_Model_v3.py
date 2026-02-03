#!/usr/bin/env python3
"""
FightingBooks Financial Model v3
Updated: 2026-02-03

NEW ASSUMPTIONS:
- 3 tiers: Free / $9.99 / $19.99 (one-time purchases)
- 3 gates × 2 choices = 8 paths per CYOA matchup
- 30 real animals (Tier 2), 47 total with dinos/fantasy (Tier 3)
- On-demand generation with caching

REALISTIC USAGE MODELING:
- Not all paths get generated (users don't explore everything)
- Caching means content generated once serves all users
- Popular matchups (lion/tiger, shark/croc) get hit early
- Long tail of rarely-used combinations
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# =============================================================================
# CORE ASSUMPTIONS
# =============================================================================

# Tier Structure
TIERS = {
    "Free": {
        "price": 0,
        "animals": 8,
        "static_matchups": 28,  # 8×7/2
        "cyoa_matchups": 1,     # Lion vs Tiger only
        "paths_per_cyoa": 8,    # 2^3 = 8
    },
    "Tier2": {
        "price": 9.99,
        "animals": 30,
        "static_matchups": 435,  # 30×29/2
        "cyoa_matchups": 435,
        "paths_per_cyoa": 8,
    },
    "Tier3": {
        "price": 19.99,
        "animals": 47,
        "static_matchups": 1081,  # 47×46/2
        "cyoa_matchups": 1081,
        "paths_per_cyoa": 8,
    }
}

# Generation Costs
COST_PER_STATIC_BOOK = 0.04  # Text + images
COST_PER_CYOA_PATH = 0.04    # Text + images for one path

# Total possible content
TOTAL_STATIC_BOOKS = 1081  # All tiers combined (Tier 3 is superset)
TOTAL_CYOA_PATHS = (1 * 8) + (435 * 8) + (1081 * 8)  # Free + T2 + T3 unique
# Simplified: Tier 3 includes all, so max unique CYOA paths = 1081 * 8 = 8,648

MAX_CYOA_PATHS = 1081 * 8  # 8,648
MAX_GENERATION_COST = (TOTAL_STATIC_BOOKS * COST_PER_STATIC_BOOK) + (MAX_CYOA_PATHS * COST_PER_CYOA_PATH)
# = $43.24 + $345.92 = $389.16 (much lower than before with 5 gates!)

# =============================================================================
# REALISTIC USAGE MODEL
# =============================================================================

"""
User Behavior Assumptions:

FREE TIER:
- Gets static books for 8 animals (28 matchups) - pre-generated
- Gets Lion vs Tiger CYOA only
- Average paths explored: 3-4 out of 8

TIER 2 ($9.99):
- Most users explore 20-40 matchups out of 435 (5-10%)
- Average CYOA paths per matchup: 3-4 out of 8 (40-50%)
- Power users might explore 100+ matchups

TIER 3 ($19.99):
- Adds dinos/fantasy - novelty drives more exploration
- Average user explores 40-80 matchups
- Power users might hit 200+

CACHING IMPACT:
- Early users generate content, later users get cached
- Popular matchups (Lion/Tiger, Shark/Croc, T-Rex/Triceratops) hit first
- ~80% of generation happens in first 20% of user base
- Long tail matchups may never get generated
"""

# Realistic content generation curve (% of max over time)
# Based on power law distribution of user interest
CONTENT_GENERATION_CURVE = {
    # Month: (static_books_%, cyoa_paths_%)
    1: (0.05, 0.02),   # Early adopters hit popular matchups
    2: (0.10, 0.05),
    3: (0.15, 0.08),
    4: (0.20, 0.12),
    5: (0.25, 0.15),
    6: (0.30, 0.18),
    7: (0.33, 0.20),
    8: (0.36, 0.22),
    9: (0.38, 0.24),
    10: (0.40, 0.25),
    11: (0.41, 0.26),
    12: (0.42, 0.27),  # Year 1: ~42% static, ~27% CYOA
    # Year 2 - slower growth, approaching asymptote
    13: (0.43, 0.28),
    14: (0.44, 0.29),
    15: (0.45, 0.30),
    16: (0.46, 0.31),
    17: (0.47, 0.32),
    18: (0.48, 0.33),
    19: (0.49, 0.34),
    20: (0.50, 0.35),
    21: (0.51, 0.36),
    22: (0.52, 0.37),
    23: (0.53, 0.38),
    24: (0.54, 0.39),  # Year 2: ~54% static, ~39% CYOA
}

# Pre-generate popular content at launch (one-time cost)
PRE_GENERATED = {
    "static_books": 28,  # Free tier animals
    "cyoa_paths": 8,     # Lion vs Tiger full CYOA
}

PRE_GEN_COST = (PRE_GENERATED["static_books"] * COST_PER_STATIC_BOOK) + \
               (PRE_GENERATED["cyoa_paths"] * COST_PER_CYOA_PATH)
# = $1.12 + $0.32 = $1.44

# =============================================================================
# FINANCIAL MODEL
# =============================================================================

def calculate_monthly_generation_cost(month, prev_static_pct, prev_cyoa_pct):
    """Calculate incremental generation cost for a month."""
    current_static_pct, current_cyoa_pct = CONTENT_GENERATION_CURVE.get(month, (0.55, 0.40))
    
    new_static = (current_static_pct - prev_static_pct) * TOTAL_STATIC_BOOKS
    new_cyoa = (current_cyoa_pct - prev_cyoa_pct) * MAX_CYOA_PATHS
    
    cost = (new_static * COST_PER_STATIC_BOOK) + (new_cyoa * COST_PER_CYOA_PATH)
    return cost, current_static_pct, current_cyoa_pct


wb = Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

def auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 18)

# =============================================================================
# SHEET 1: ASSUMPTIONS
# =============================================================================
ws_assumptions = wb.active
ws_assumptions.title = "Assumptions"

assumptions_data = [
    ["FIGHTINGBOOKS FINANCIAL MODEL v3"],
    ["Updated: 2026-02-03"],
    [],
    ["TIER STRUCTURE"],
    ["Tier", "Price", "Animals", "Static Books", "CYOA Matchups", "Paths/CYOA", "Total Paths"],
    ["Free", "$0", 8, 28, 1, 8, 8],
    ["Tier 2", "$9.99", 30, 435, 435, 8, 3480],
    ["Tier 3", "$19.99", 47, 1081, 1081, 8, 8648],
    [],
    ["GENERATION COSTS"],
    ["Item", "Cost"],
    ["Static book (text + images)", "$0.04"],
    ["CYOA path (text + images)", "$0.04"],
    [],
    ["MAXIMUM CONTENT"],
    ["Item", "Count", "Max Cost"],
    ["Static books (all tiers)", 1081, f"${1081 * 0.04:.2f}"],
    ["CYOA paths (all tiers)", 8648, f"${8648 * 0.04:.2f}"],
    ["TOTAL MAX", "", f"${(1081 + 8648) * 0.04:.2f}"],
    [],
    ["PRE-GENERATED AT LAUNCH"],
    ["Item", "Count", "Cost"],
    ["Free tier static books", 28, f"${28 * 0.04:.2f}"],
    ["Lion vs Tiger CYOA (all paths)", 8, f"${8 * 0.04:.2f}"],
    ["TOTAL PRE-GEN", "", f"${PRE_GEN_COST:.2f}"],
    [],
    ["REALISTIC USAGE (Year 1 End)"],
    ["Static books generated", "~42%", f"~{int(1081 * 0.42)} books"],
    ["CYOA paths generated", "~27%", f"~{int(8648 * 0.27)} paths"],
    ["Estimated actual cost", "", f"~${(1081 * 0.42 * 0.04) + (8648 * 0.27 * 0.04):.2f}"],
    [],
    ["REALISTIC USAGE (Year 2 End)"],
    ["Static books generated", "~54%", f"~{int(1081 * 0.54)} books"],
    ["CYOA paths generated", "~39%", f"~{int(8648 * 0.39)} paths"],
    ["Estimated actual cost", "", f"~${(1081 * 0.54 * 0.04) + (8648 * 0.39 * 0.04):.2f}"],
    [],
    ["PAYMENT PROCESSING"],
    ["Stripe fee", "2.9% + $0.30"],
    ["Tier 2 net", f"${9.99 * 0.971 - 0.30:.2f}"],
    ["Tier 3 net", f"${19.99 * 0.971 - 0.30:.2f}"],
    [],
    ["OTHER COSTS"],
    ["Domain (annual)", "$12"],
    ["Vercel hosting", "$0 (free tier)"],
    ["Marketing budget (monthly)", "$50-150"],
]

for row in assumptions_data:
    ws_assumptions.append(row)

auto_width(ws_assumptions)

# =============================================================================
# SHEET 2: CONSERVATIVE SCENARIO
# =============================================================================
ws_cons = wb.create_sheet("Conservative")

# Conservative assumptions
# - Slow organic growth
# - 8% Tier 2 conversion, 3% Tier 3 conversion
# - Seasonal dips in summer/holidays

cons_new_free = [100, 120, 150, 180, 220, 250, 280, 300, 320, 280, 220, 180,  # Y1: 2,600
                 200, 240, 300, 350, 400, 420, 450, 480, 500, 400, 320, 260]  # Y2: 4,320

cons_t2_rate = 0.08  # 8% of free users buy Tier 2
cons_t3_rate = 0.03  # 3% of free users buy Tier 3
cons_marketing = [50, 50, 75, 75, 100, 100, 100, 100, 100, 75, 50, 50,  # Y1: $925
                  75, 75, 100, 100, 125, 125, 125, 125, 125, 100, 75, 75]  # Y2: $1,225

# Headers
headers = ["Metric"] + [f"M{i}" for i in range(1, 25)] + ["Y1", "Y2", "Total"]
ws_cons.append(["CONSERVATIVE SCENARIO - 24 Month Projections"])
ws_cons.append([])
ws_cons.append(headers)

# Calculate data
cons_t2_sales = [int(f * cons_t2_rate) for f in cons_new_free]
cons_t3_sales = [int(f * cons_t3_rate) for f in cons_new_free]
cons_t2_rev = [s * (9.99 * 0.971 - 0.30) for s in cons_t2_sales]  # Net after Stripe
cons_t3_rev = [s * (19.99 * 0.971 - 0.30) for s in cons_t3_sales]
cons_total_rev = [t2 + t3 for t2, t3 in zip(cons_t2_rev, cons_t3_rev)]

# Generation costs (realistic curve)
cons_gen_costs = []
prev_static, prev_cyoa = 0, 0
for m in range(1, 25):
    cost, prev_static, prev_cyoa = calculate_monthly_generation_cost(m, prev_static, prev_cyoa)
    cons_gen_costs.append(cost)

# Add pre-gen cost to month 1
cons_gen_costs[0] += PRE_GEN_COST

# Total costs
cons_total_costs = [gen + mkt for gen, mkt in zip(cons_gen_costs, cons_marketing)]

# Profit
cons_profit = [rev - cost for rev, cost in zip(cons_total_rev, cons_total_costs)]

# Add rows
rows = [
    ("New Free Users", cons_new_free),
    ("Tier 2 Sales", cons_t2_sales),
    ("Tier 3 Sales", cons_t3_sales),
    ("", []),
    ("Tier 2 Revenue", cons_t2_rev),
    ("Tier 3 Revenue", cons_t3_rev),
    ("Total Revenue", cons_total_rev),
    ("", []),
    ("Generation Costs", cons_gen_costs),
    ("Marketing Costs", cons_marketing),
    ("Total Costs", cons_total_costs),
    ("", []),
    ("Monthly Profit", cons_profit),
]

for label, data in rows:
    if data:
        y1 = sum(data[:12])
        y2 = sum(data[12:])
        total = y1 + y2
        formatted = [round(d, 2) if isinstance(d, float) else d for d in data]
        ws_cons.append([label] + formatted + [round(y1, 2), round(y2, 2), round(total, 2)])
    else:
        ws_cons.append([label])

# Cumulative profit
cum_profit = []
running = 0
for p in cons_profit:
    running += p
    cum_profit.append(running)
ws_cons.append(["Cumulative Profit"] + [round(c, 2) for c in cum_profit] + ["", "", round(cum_profit[-1], 2)])

auto_width(ws_cons)

# =============================================================================
# SHEET 3: OPTIMISTIC SCENARIO
# =============================================================================
ws_opt = wb.create_sheet("Optimistic")

# Optimistic assumptions
# - Viral moments, strong word-of-mouth
# - 12% Tier 2 conversion, 5% Tier 3 conversion
# - Higher marketing spend

opt_new_free = [200, 300, 450, 600, 800, 1000, 1200, 1400, 1600, 1400, 1000, 800,  # Y1: 10,750
                900, 1100, 1400, 1700, 2000, 2200, 2400, 2600, 2800, 2400, 1800, 1400]  # Y2: 22,700

opt_t2_rate = 0.12  # 12% conversion
opt_t3_rate = 0.05  # 5% conversion
opt_marketing = [100, 100, 150, 150, 200, 200, 200, 200, 200, 150, 100, 100,  # Y1: $1,850
                 150, 150, 200, 200, 250, 250, 250, 250, 250, 200, 150, 150]  # Y2: $2,450

ws_opt.append(["OPTIMISTIC SCENARIO - 24 Month Projections"])
ws_opt.append([])
ws_opt.append(headers)

# Calculate
opt_t2_sales = [int(f * opt_t2_rate) for f in opt_new_free]
opt_t3_sales = [int(f * opt_t3_rate) for f in opt_new_free]
opt_t2_rev = [s * (9.99 * 0.971 - 0.30) for s in opt_t2_sales]
opt_t3_rev = [s * (19.99 * 0.971 - 0.30) for s in opt_t3_sales]
opt_total_rev = [t2 + t3 for t2, t3 in zip(opt_t2_rev, opt_t3_rev)]

# Generation costs (same curve, just reaches it faster with more users)
# In optimistic, we hit the ceiling faster
opt_gen_curve = {
    1: (0.10, 0.05),
    2: (0.20, 0.12),
    3: (0.30, 0.20),
    4: (0.40, 0.28),
    5: (0.48, 0.35),
    6: (0.54, 0.40),
    7: (0.58, 0.44),
    8: (0.61, 0.47),
    9: (0.63, 0.49),
    10: (0.65, 0.51),
    11: (0.66, 0.52),
    12: (0.67, 0.53),
    13: (0.68, 0.54),
    14: (0.69, 0.55),
    15: (0.70, 0.56),
    16: (0.71, 0.57),
    17: (0.72, 0.58),
    18: (0.73, 0.59),
    19: (0.74, 0.60),
    20: (0.75, 0.61),
    21: (0.76, 0.62),
    22: (0.77, 0.63),
    23: (0.78, 0.64),
    24: (0.79, 0.65),
}

opt_gen_costs = []
prev_static, prev_cyoa = 0, 0
for m in range(1, 25):
    static_pct, cyoa_pct = opt_gen_curve.get(m, (0.80, 0.66))
    new_static = (static_pct - prev_static) * TOTAL_STATIC_BOOKS
    new_cyoa = (cyoa_pct - prev_cyoa) * MAX_CYOA_PATHS
    cost = (new_static * COST_PER_STATIC_BOOK) + (new_cyoa * COST_PER_CYOA_PATH)
    opt_gen_costs.append(cost)
    prev_static, prev_cyoa = static_pct, cyoa_pct

opt_gen_costs[0] += PRE_GEN_COST

opt_total_costs = [gen + mkt for gen, mkt in zip(opt_gen_costs, opt_marketing)]
opt_profit = [rev - cost for rev, cost in zip(opt_total_rev, opt_total_costs)]

# Add rows
opt_rows = [
    ("New Free Users", opt_new_free),
    ("Tier 2 Sales", opt_t2_sales),
    ("Tier 3 Sales", opt_t3_sales),
    ("", []),
    ("Tier 2 Revenue", opt_t2_rev),
    ("Tier 3 Revenue", opt_t3_rev),
    ("Total Revenue", opt_total_rev),
    ("", []),
    ("Generation Costs", opt_gen_costs),
    ("Marketing Costs", opt_marketing),
    ("Total Costs", opt_total_costs),
    ("", []),
    ("Monthly Profit", opt_profit),
]

for label, data in opt_rows:
    if data:
        y1 = sum(data[:12])
        y2 = sum(data[12:])
        total = y1 + y2
        formatted = [round(d, 2) if isinstance(d, float) else d for d in data]
        ws_opt.append([label] + formatted + [round(y1, 2), round(y2, 2), round(total, 2)])
    else:
        ws_opt.append([label])

# Cumulative
opt_cum_profit = []
running = 0
for p in opt_profit:
    running += p
    opt_cum_profit.append(running)
ws_opt.append(["Cumulative Profit"] + [round(c, 2) for c in opt_cum_profit] + ["", "", round(opt_cum_profit[-1], 2)])

auto_width(ws_opt)

# =============================================================================
# SHEET 4: SUMMARY
# =============================================================================
ws_summary = wb.create_sheet("Summary")

summary_data = [
    ["FIGHTINGBOOKS 24-MONTH SUMMARY"],
    [],
    ["SCENARIO COMPARISON"],
    ["Metric", "Conservative", "Optimistic"],
    ["Total Free Users (2Y)", sum(cons_new_free), sum(opt_new_free)],
    ["Total Tier 2 Sales", sum(cons_t2_sales), sum(opt_t2_sales)],
    ["Total Tier 3 Sales", sum(cons_t3_sales), sum(opt_t3_sales)],
    ["Total Paid Customers", sum(cons_t2_sales) + sum(cons_t3_sales), sum(opt_t2_sales) + sum(opt_t3_sales)],
    [],
    ["Total Revenue", f"${sum(cons_total_rev):,.2f}", f"${sum(opt_total_rev):,.2f}"],
    ["Total Costs", f"${sum(cons_total_costs):,.2f}", f"${sum(opt_total_costs):,.2f}"],
    ["Total Profit", f"${sum(cons_profit):,.2f}", f"${sum(opt_profit):,.2f}"],
    [],
    ["Break-even Month", next((i+1 for i, c in enumerate(cum_profit) if c > 0), "N/A"), 
     next((i+1 for i, c in enumerate(opt_cum_profit) if c > 0), "N/A")],
    [],
    ["CONTENT GENERATION"],
    ["", "Conservative", "Optimistic"],
    ["Year 1 Gen Cost", f"${sum(cons_gen_costs[:12]):,.2f}", f"${sum(opt_gen_costs[:12]):,.2f}"],
    ["Year 2 Gen Cost", f"${sum(cons_gen_costs[12:]):,.2f}", f"${sum(opt_gen_costs[12:]):,.2f}"],
    ["Total Gen Cost", f"${sum(cons_gen_costs):,.2f}", f"${sum(opt_gen_costs):,.2f}"],
    ["% of Max Content", "~54%", "~79%"],
    [],
    ["KEY METRICS"],
    ["Metric", "Conservative", "Optimistic"],
    ["Conversion Rate (T2)", f"{cons_t2_rate*100:.0f}%", f"{opt_t2_rate*100:.0f}%"],
    ["Conversion Rate (T3)", f"{cons_t3_rate*100:.0f}%", f"{opt_t3_rate*100:.0f}%"],
    ["Avg Revenue/Free User", f"${sum(cons_total_rev)/sum(cons_new_free):.2f}", 
     f"${sum(opt_total_rev)/sum(opt_new_free):.2f}"],
    ["Marketing ROI", f"{sum(cons_total_rev)/sum(cons_marketing):.1f}x", 
     f"{sum(opt_total_rev)/sum(opt_marketing):.1f}x"],
    [],
    ["RISK ANALYSIS"],
    ["Worst Case Max Cost", f"${MAX_GENERATION_COST:.2f}", "(if 100% content generated)"],
    ["Realistic Max Cost", f"${MAX_GENERATION_COST * 0.6:.2f}", "(if 60% content generated)"],
    ["Break-even Sales (worst)", f"{int(MAX_GENERATION_COST / 9.40) + 1}", "Tier 2 sales"],
    ["Break-even Sales (realistic)", f"{int(MAX_GENERATION_COST * 0.3 / 9.40) + 1}", "Tier 2 sales"],
]

for row in summary_data:
    ws_summary.append(row)

auto_width(ws_summary)

# =============================================================================
# SAVE
# =============================================================================
output_path = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(output_path, "FightingBooks_Financial_Model_v3.xlsx")
wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")

# Print summary
print("\n" + "="*60)
print("FIGHTINGBOOKS v3 FINANCIAL SUMMARY")
print("="*60)
print(f"\nTIER STRUCTURE:")
print(f"  Free:   8 animals, 28 static books, 1 CYOA (Lion vs Tiger)")
print(f"  Tier 2: 30 animals, 435 matchups, unlimited CYOA - $9.99")
print(f"  Tier 3: 47 animals, 1081 matchups, unlimited CYOA - $19.99")
print(f"\nCYOA: 3 gates × 2 choices = 8 paths per matchup")
print(f"\nMAX CONTENT:")
print(f"  Static books: 1,081")
print(f"  CYOA paths:   8,648")
print(f"  Max cost:     ${MAX_GENERATION_COST:.2f}")
print(f"\nREALISTIC USAGE (Year 2):")
print(f"  Conservative: ~54% content, ~${sum(cons_gen_costs):.2f} gen cost")
print(f"  Optimistic:   ~79% content, ~${sum(opt_gen_costs):.2f} gen cost")
print(f"\n24-MONTH PROJECTIONS:")
print(f"  Conservative: ${sum(cons_total_rev):,.2f} revenue → ${sum(cons_profit):,.2f} profit")
print(f"  Optimistic:   ${sum(opt_total_rev):,.2f} revenue → ${sum(opt_profit):,.2f} profit")
print(f"\nBREAK-EVEN:")
print(f"  Conservative: Month {next((i+1 for i, c in enumerate(cum_profit) if c > 0), 'N/A')}")
print(f"  Optimistic:   Month {next((i+1 for i, c in enumerate(opt_cum_profit) if c > 0), 'N/A')}")
