#!/usr/bin/env python3
"""
FightingBooks Financial Model v2
Month-on-month projections for 2 years, Conservative and Optimistic scenarios
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
money_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
money_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

def auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 15)

# ============================================================
# CONSERVATIVE SCENARIO - 24 months
# ============================================================
ws1 = wb.active
ws1.title = "Conservative"

# Monthly growth assumptions - Conservative
# Starts slow, builds momentum, seasonal dips
cons_free = [150, 180, 220, 280, 350, 400, 450, 500, 550, 450, 350, 300,  # Y1
             350, 400, 500, 600, 700, 750, 800, 850, 900, 700, 500, 400]  # Y2
cons_premium_rate = [0.08, 0.08, 0.09, 0.09, 0.10, 0.10, 0.10, 0.11, 0.11, 0.10, 0.09, 0.08,
                     0.10, 0.10, 0.11, 0.11, 0.12, 0.12, 0.12, 0.12, 0.12, 0.11, 0.10, 0.09]
cons_ultimate_rate = [0.03, 0.03, 0.04, 0.04, 0.05, 0.05, 0.05, 0.05, 0.06, 0.05, 0.04, 0.03,
                      0.05, 0.05, 0.06, 0.06, 0.06, 0.07, 0.07, 0.07, 0.07, 0.06, 0.05, 0.04]
cons_cyoa_plays_per_ultimate = [30, 35, 40, 45, 50, 50, 50, 45, 45, 40, 35, 30,
                                 35, 40, 45, 50, 50, 55, 55, 50, 50, 45, 40, 35]
cons_cache_hit = [0.15, 0.25, 0.35, 0.45, 0.52, 0.58, 0.63, 0.68, 0.72, 0.75, 0.78, 0.80,
                  0.82, 0.84, 0.85, 0.86, 0.87, 0.88, 0.89, 0.90, 0.90, 0.91, 0.91, 0.92]

# Marketing spend by month - Conservative
cons_marketing = [50, 50, 100, 100, 150, 150, 150, 150, 150, 100, 100, 100,  # Y1 = $1,350
                  100, 100, 150, 150, 200, 200, 200, 200, 200, 150, 100, 100]  # Y2 = $1,850

# Headers
headers = ["Metric"] + [f"M{i}" for i in range(1, 25)] + ["Y1 Total", "Y2 Total", "2Y Total"]
ws1.append(["CONSERVATIVE SCENARIO - 24 Month Projections"])
ws1.append([])
ws1.append(headers)

# Calculate monthly data
rows_data = []

# New Free Users
rows_data.append(["New Free Users"] + cons_free + [sum(cons_free[:12]), sum(cons_free[12:]), sum(cons_free)])

# Cumulative Free Users
cum_free = []
total = 0
for f in cons_free:
    total += f
    cum_free.append(total)
rows_data.append(["Cumulative Free"] + cum_free + [cum_free[11], cum_free[23], cum_free[23]])

# New Premium Users
cons_premium = [int(cons_free[i] * cons_premium_rate[i]) for i in range(24)]
rows_data.append(["New Premium Users"] + cons_premium + [sum(cons_premium[:12]), sum(cons_premium[12:]), sum(cons_premium)])

# New Ultimate Users
cons_ultimate = [int(cons_free[i] * cons_ultimate_rate[i]) for i in range(24)]
rows_data.append(["New Ultimate Users"] + cons_ultimate + [sum(cons_ultimate[:12]), sum(cons_ultimate[12:]), sum(cons_ultimate)])

# Cumulative Ultimate (for CYOA calc)
cum_ultimate = []
total = 0
for u in cons_ultimate:
    total += u
    cum_ultimate.append(total)

rows_data.append([])  # Spacer

# Revenue
premium_rev = [p * 9.99 for p in cons_premium]
rows_data.append(["Premium Revenue ($9.99)"] + [round(r, 2) for r in premium_rev] + 
                 [round(sum(premium_rev[:12]), 2), round(sum(premium_rev[12:]), 2), round(sum(premium_rev), 2)])

ultimate_rev = [u * 19.99 for u in cons_ultimate]
rows_data.append(["Ultimate Revenue ($19.99)"] + [round(r, 2) for r in ultimate_rev] + 
                 [round(sum(ultimate_rev[:12]), 2), round(sum(ultimate_rev[12:]), 2), round(sum(ultimate_rev), 2)])

gross_rev = [premium_rev[i] + ultimate_rev[i] for i in range(24)]
rows_data.append(["GROSS REVENUE"] + [round(r, 2) for r in gross_rev] + 
                 [round(sum(gross_rev[:12]), 2), round(sum(gross_rev[12:]), 2), round(sum(gross_rev), 2)])

# Refunds (5%)
refunds = [r * 0.05 for r in gross_rev]
rows_data.append(["Refunds (5%)"] + [round(r, 2) for r in refunds] + 
                 [round(sum(refunds[:12]), 2), round(sum(refunds[12:]), 2), round(sum(refunds), 2)])

net_rev = [gross_rev[i] - refunds[i] for i in range(24)]
rows_data.append(["NET REVENUE"] + [round(r, 2) for r in net_rev] + 
                 [round(sum(net_rev[:12]), 2), round(sum(net_rev[12:]), 2), round(sum(net_rev), 2)])

rows_data.append([])  # Spacer

# Costs
# Book generation (new books only, assume 50% are new matchups initially, decreasing)
new_book_rate = [0.6, 0.5, 0.4, 0.35, 0.3, 0.25, 0.22, 0.20, 0.18, 0.16, 0.14, 0.12,
                 0.12, 0.11, 0.10, 0.10, 0.09, 0.09, 0.08, 0.08, 0.07, 0.07, 0.06, 0.06]
books_generated = [int(cons_free[i] * new_book_rate[i]) for i in range(24)]
book_gen_cost = [b * 0.20 for b in books_generated]
rows_data.append(["Book Gen Cost ($0.20/new)"] + [round(c, 2) for c in book_gen_cost] + 
                 [round(sum(book_gen_cost[:12]), 2), round(sum(book_gen_cost[12:]), 2), round(sum(book_gen_cost), 2)])

# CYOA costs
cyoa_plays = [cum_ultimate[i] * cons_cyoa_plays_per_ultimate[i] for i in range(24)]
cyoa_misses = [cyoa_plays[i] * (1 - cons_cache_hit[i]) for i in range(24)]
cyoa_cost = [m * 0.13 for m in cyoa_misses]
rows_data.append(["CYOA Cost ($0.13/miss)"] + [round(c, 2) for c in cyoa_cost] + 
                 [round(sum(cyoa_cost[:12]), 2), round(sum(cyoa_cost[12:]), 2), round(sum(cyoa_cost), 2)])

# Storage (grows with cache)
storage_gb = [1, 1.5, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
              12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]
storage_cost = [max(0, (s - 10) * 0.15) for s in storage_gb]  # First 10GB free
rows_data.append(["Storage Cost"] + [round(c, 2) for c in storage_cost] + 
                 [round(sum(storage_cost[:12]), 2), round(sum(storage_cost[12:]), 2), round(sum(storage_cost), 2)])

# Marketing
rows_data.append(["Marketing"] + cons_marketing + [sum(cons_marketing[:12]), sum(cons_marketing[12:]), sum(cons_marketing)])

# Domain (month 1 only)
domain_cost = [12] + [0] * 11 + [12] + [0] * 11
rows_data.append(["Domain"] + domain_cost + [12, 12, 24])

# Total Costs
total_costs = [book_gen_cost[i] + cyoa_cost[i] + storage_cost[i] + cons_marketing[i] + domain_cost[i] for i in range(24)]
rows_data.append(["TOTAL COSTS"] + [round(c, 2) for c in total_costs] + 
                 [round(sum(total_costs[:12]), 2), round(sum(total_costs[12:]), 2), round(sum(total_costs), 2)])

rows_data.append([])  # Spacer

# Profit
monthly_profit = [net_rev[i] - total_costs[i] for i in range(24)]
rows_data.append(["MONTHLY PROFIT"] + [round(p, 2) for p in monthly_profit] + 
                 [round(sum(monthly_profit[:12]), 2), round(sum(monthly_profit[12:]), 2), round(sum(monthly_profit), 2)])

# Cumulative Profit
cum_profit = []
total = 0
for p in monthly_profit:
    total += p
    cum_profit.append(round(total, 2))
rows_data.append(["CUMULATIVE PROFIT"] + cum_profit + [cum_profit[11], cum_profit[23], cum_profit[23]])

# Margin
margins = [monthly_profit[i] / net_rev[i] * 100 if net_rev[i] > 0 else 0 for i in range(24)]
rows_data.append(["Profit Margin %"] + [f"{m:.1f}%" for m in margins] + ["", "", ""])

# Write all rows
for row_data in rows_data:
    ws1.append(row_data)

style_header(ws1, 3, 1, 28)
apply_borders(ws1, 3, ws1.max_row, 1, 28)
auto_width(ws1)

# ============================================================
# OPTIMISTIC SCENARIO - 24 months
# ============================================================
ws2 = wb.create_sheet("Optimistic")

# Optimistic assumptions - faster growth, higher conversion
opt_free = [400, 600, 900, 1300, 1800, 2200, 2500, 2800, 3000, 2500, 2000, 1500,  # Y1
            2000, 2500, 3500, 4500, 5500, 6000, 6500, 7000, 7000, 5500, 4000, 3000]  # Y2
opt_premium_rate = [0.10, 0.11, 0.12, 0.13, 0.14, 0.14, 0.15, 0.15, 0.15, 0.14, 0.12, 0.10,
                    0.12, 0.13, 0.14, 0.15, 0.15, 0.16, 0.16, 0.16, 0.16, 0.15, 0.13, 0.11]
opt_ultimate_rate = [0.05, 0.05, 0.06, 0.07, 0.07, 0.08, 0.08, 0.08, 0.08, 0.07, 0.06, 0.05,
                     0.06, 0.07, 0.07, 0.08, 0.08, 0.09, 0.09, 0.09, 0.09, 0.08, 0.07, 0.06]
opt_cyoa_plays = [35, 40, 45, 50, 55, 55, 55, 50, 50, 45, 40, 35,
                  40, 45, 50, 55, 55, 60, 60, 55, 55, 50, 45, 40]
opt_cache_hit = [0.10, 0.20, 0.30, 0.40, 0.48, 0.55, 0.60, 0.65, 0.70, 0.74, 0.77, 0.80,
                 0.82, 0.84, 0.86, 0.87, 0.88, 0.89, 0.90, 0.91, 0.91, 0.92, 0.92, 0.93]
opt_marketing = [100, 150, 200, 250, 300, 350, 350, 350, 350, 250, 200, 150,  # Y1 = $3,000
                 200, 250, 350, 400, 500, 500, 500, 500, 500, 400, 300, 200]  # Y2 = $4,600

# Headers
ws2.append(["OPTIMISTIC SCENARIO - 24 Month Projections"])
ws2.append([])
ws2.append(headers)

# Calculate optimistic data
rows_data2 = []

rows_data2.append(["New Free Users"] + opt_free + [sum(opt_free[:12]), sum(opt_free[12:]), sum(opt_free)])

cum_free2 = []
total = 0
for f in opt_free:
    total += f
    cum_free2.append(total)
rows_data2.append(["Cumulative Free"] + cum_free2 + [cum_free2[11], cum_free2[23], cum_free2[23]])

opt_premium = [int(opt_free[i] * opt_premium_rate[i]) for i in range(24)]
rows_data2.append(["New Premium Users"] + opt_premium + [sum(opt_premium[:12]), sum(opt_premium[12:]), sum(opt_premium)])

opt_ultimate = [int(opt_free[i] * opt_ultimate_rate[i]) for i in range(24)]
rows_data2.append(["New Ultimate Users"] + opt_ultimate + [sum(opt_ultimate[:12]), sum(opt_ultimate[12:]), sum(opt_ultimate)])

cum_ultimate2 = []
total = 0
for u in opt_ultimate:
    total += u
    cum_ultimate2.append(total)

rows_data2.append([])

premium_rev2 = [p * 9.99 for p in opt_premium]
rows_data2.append(["Premium Revenue ($9.99)"] + [round(r, 2) for r in premium_rev2] + 
                  [round(sum(premium_rev2[:12]), 2), round(sum(premium_rev2[12:]), 2), round(sum(premium_rev2), 2)])

ultimate_rev2 = [u * 19.99 for u in opt_ultimate]
rows_data2.append(["Ultimate Revenue ($19.99)"] + [round(r, 2) for r in ultimate_rev2] + 
                  [round(sum(ultimate_rev2[:12]), 2), round(sum(ultimate_rev2[12:]), 2), round(sum(ultimate_rev2), 2)])

gross_rev2 = [premium_rev2[i] + ultimate_rev2[i] for i in range(24)]
rows_data2.append(["GROSS REVENUE"] + [round(r, 2) for r in gross_rev2] + 
                  [round(sum(gross_rev2[:12]), 2), round(sum(gross_rev2[12:]), 2), round(sum(gross_rev2), 2)])

refunds2 = [r * 0.05 for r in gross_rev2]
rows_data2.append(["Refunds (5%)"] + [round(r, 2) for r in refunds2] + 
                  [round(sum(refunds2[:12]), 2), round(sum(refunds2[12:]), 2), round(sum(refunds2), 2)])

net_rev2 = [gross_rev2[i] - refunds2[i] for i in range(24)]
rows_data2.append(["NET REVENUE"] + [round(r, 2) for r in net_rev2] + 
                  [round(sum(net_rev2[:12]), 2), round(sum(net_rev2[12:]), 2), round(sum(net_rev2), 2)])

rows_data2.append([])

# Costs - Optimistic
new_book_rate2 = [0.5, 0.4, 0.32, 0.26, 0.22, 0.18, 0.15, 0.13, 0.11, 0.10, 0.09, 0.08,
                  0.08, 0.07, 0.07, 0.06, 0.06, 0.05, 0.05, 0.05, 0.04, 0.04, 0.04, 0.04]
books_generated2 = [int(opt_free[i] * new_book_rate2[i]) for i in range(24)]
book_gen_cost2 = [b * 0.20 for b in books_generated2]
rows_data2.append(["Book Gen Cost ($0.20/new)"] + [round(c, 2) for c in book_gen_cost2] + 
                  [round(sum(book_gen_cost2[:12]), 2), round(sum(book_gen_cost2[12:]), 2), round(sum(book_gen_cost2), 2)])

cyoa_plays2 = [cum_ultimate2[i] * opt_cyoa_plays[i] for i in range(24)]
cyoa_misses2 = [cyoa_plays2[i] * (1 - opt_cache_hit[i]) for i in range(24)]
cyoa_cost2 = [m * 0.13 for m in cyoa_misses2]
rows_data2.append(["CYOA Cost ($0.13/miss)"] + [round(c, 2) for c in cyoa_cost2] + 
                  [round(sum(cyoa_cost2[:12]), 2), round(sum(cyoa_cost2[12:]), 2), round(sum(cyoa_cost2), 2)])

storage_gb2 = [2, 4, 7, 11, 16, 21, 26, 31, 36, 40, 43, 46,
               50, 54, 58, 62, 66, 70, 74, 78, 82, 85, 88, 90]
storage_cost2 = [max(0, (s - 10) * 0.15) for s in storage_gb2]
rows_data2.append(["Storage Cost"] + [round(c, 2) for c in storage_cost2] + 
                  [round(sum(storage_cost2[:12]), 2), round(sum(storage_cost2[12:]), 2), round(sum(storage_cost2), 2)])

rows_data2.append(["Marketing"] + opt_marketing + [sum(opt_marketing[:12]), sum(opt_marketing[12:]), sum(opt_marketing)])

domain_cost2 = [12] + [0] * 11 + [12] + [0] * 11
rows_data2.append(["Domain"] + domain_cost2 + [12, 12, 24])

total_costs2 = [book_gen_cost2[i] + cyoa_cost2[i] + storage_cost2[i] + opt_marketing[i] + domain_cost2[i] for i in range(24)]
rows_data2.append(["TOTAL COSTS"] + [round(c, 2) for c in total_costs2] + 
                  [round(sum(total_costs2[:12]), 2), round(sum(total_costs2[12:]), 2), round(sum(total_costs2), 2)])

rows_data2.append([])

monthly_profit2 = [net_rev2[i] - total_costs2[i] for i in range(24)]
rows_data2.append(["MONTHLY PROFIT"] + [round(p, 2) for p in monthly_profit2] + 
                  [round(sum(monthly_profit2[:12]), 2), round(sum(monthly_profit2[12:]), 2), round(sum(monthly_profit2), 2)])

cum_profit2 = []
total = 0
for p in monthly_profit2:
    total += p
    cum_profit2.append(round(total, 2))
rows_data2.append(["CUMULATIVE PROFIT"] + cum_profit2 + [cum_profit2[11], cum_profit2[23], cum_profit2[23]])

margins2 = [monthly_profit2[i] / net_rev2[i] * 100 if net_rev2[i] > 0 else 0 for i in range(24)]
rows_data2.append(["Profit Margin %"] + [f"{m:.1f}%" for m in margins2] + ["", "", ""])

for row_data in rows_data2:
    ws2.append(row_data)

style_header(ws2, 3, 1, 28)
apply_borders(ws2, 3, ws2.max_row, 1, 28)
auto_width(ws2)

# ============================================================
# SUMMARY COMPARISON
# ============================================================
ws3 = wb.create_sheet("Summary")

ws3.append(["SCENARIO COMPARISON SUMMARY"])
ws3.append([])
ws3.append(["Metric", "Conservative Y1", "Conservative Y2", "Optimistic Y1", "Optimistic Y2"])
ws3.append(["New Free Users", sum(cons_free[:12]), sum(cons_free[12:]), sum(opt_free[:12]), sum(opt_free[12:])])
ws3.append(["New Premium Users", sum(cons_premium[:12]), sum(cons_premium[12:]), sum(opt_premium[:12]), sum(opt_premium[12:])])
ws3.append(["New Ultimate Users", sum(cons_ultimate[:12]), sum(cons_ultimate[12:]), sum(opt_ultimate[:12]), sum(opt_ultimate[12:])])
ws3.append([])
ws3.append(["Gross Revenue", round(sum(gross_rev[:12]), 2), round(sum(gross_rev[12:]), 2), 
            round(sum(gross_rev2[:12]), 2), round(sum(gross_rev2[12:]), 2)])
ws3.append(["Total Costs", round(sum(total_costs[:12]), 2), round(sum(total_costs[12:]), 2),
            round(sum(total_costs2[:12]), 2), round(sum(total_costs2[12:]), 2)])
ws3.append(["Net Profit", round(sum(monthly_profit[:12]), 2), round(sum(monthly_profit[12:]), 2),
            round(sum(monthly_profit2[:12]), 2), round(sum(monthly_profit2[12:]), 2)])
ws3.append([])
ws3.append(["2-Year Totals", "Conservative", "", "Optimistic", ""])
ws3.append(["Total Revenue", round(sum(gross_rev), 2), "", round(sum(gross_rev2), 2), ""])
ws3.append(["Total Costs", round(sum(total_costs), 2), "", round(sum(total_costs2), 2), ""])
ws3.append(["Total Profit", round(sum(monthly_profit), 2), "", round(sum(monthly_profit2), 2), ""])
ws3.append(["Avg Margin", f"{sum(monthly_profit)/sum(net_rev)*100:.1f}%", "", f"{sum(monthly_profit2)/sum(net_rev2)*100:.1f}%", ""])

style_header(ws3, 3, 1, 5)
style_header(ws3, 12, 1, 5)
apply_borders(ws3, 3, ws3.max_row, 1, 5)
auto_width(ws3)

# ============================================================
# ASSUMPTIONS
# ============================================================
ws4 = wb.create_sheet("Assumptions")

assumptions = [
    ["KEY ASSUMPTIONS"],
    [],
    ["Pricing"],
    ["Premium Tier", "$9.99", "One-time"],
    ["Ultimate Tier", "$19.99", "One-time"],
    [],
    ["Conversion Rates (Conservative)"],
    ["Free → Premium", "8-11%", "Varies by month"],
    ["Free → Ultimate", "3-6%", "Varies by month"],
    [],
    ["Conversion Rates (Optimistic)"],
    ["Free → Premium", "10-16%", "Varies by month"],
    ["Free → Ultimate", "5-9%", "Varies by month"],
    [],
    ["Variable Costs"],
    ["Book Generation", "$0.20", "Per new book (images + text)"],
    ["CYOA Play (cache miss)", "$0.13", "Per uncached path"],
    ["Storage (over 10GB)", "$0.15/GB", "Monthly"],
    [],
    ["Fixed Costs"],
    ["Domain", "$12", "Annual"],
    ["Vercel Hosting", "$0", "Free tier"],
    [],
    ["Cache Assumptions"],
    ["Initial Cache Hit Rate", "15-20%", "Month 1"],
    ["Mature Cache Hit Rate", "90-93%", "By month 24"],
    [],
    ["Other"],
    ["Refund Rate", "5%", "Of gross revenue"],
    ["Seasonality", "Yes", "Dips in winter months"],
]

for row in assumptions:
    ws4.append(row)

style_header(ws4, 1, 1, 3)
style_header(ws4, 3, 1, 3)
style_header(ws4, 7, 1, 3)
style_header(ws4, 11, 1, 3)
style_header(ws4, 15, 1, 3)
style_header(ws4, 20, 1, 3)
style_header(ws4, 24, 1, 3)
style_header(ws4, 28, 1, 3)
auto_width(ws4)

# Save
output_path = "/home/ubuntu/clawd/projects/fightingbooks-web/business-plan/FightingBooks_Financial_Model_24mo.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")

# Print summary
print("\n=== SUMMARY ===")
print(f"\nConservative (2 years):")
print(f"  Revenue: ${sum(gross_rev):,.2f}")
print(f"  Costs: ${sum(total_costs):,.2f}")
print(f"  Profit: ${sum(monthly_profit):,.2f}")

print(f"\nOptimistic (2 years):")
print(f"  Revenue: ${sum(gross_rev2):,.2f}")
print(f"  Costs: ${sum(total_costs2):,.2f}")
print(f"  Profit: ${sum(monthly_profit2):,.2f}")
