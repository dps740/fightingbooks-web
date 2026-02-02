#!/usr/bin/env python3
"""
FightingBooks Financial Model Generator
Creates detailed Excel workbook with business plan financials
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
money_format = '"$"#,##0.00'
percent_format = '0.0%'
number_format = '#,##0'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 40)

# ============================================================
# SHEET 1: Executive Summary
# ============================================================
ws1 = wb.active
ws1.title = "Executive Summary"

summary_data = [
    ["FIGHTINGBOOKS BUSINESS PLAN", "", ""],
    ["", "", ""],
    ["Executive Summary", "", ""],
    ["", "", ""],
    ["Product", "AI-generated 'Who Would Win?' style battle books for kids", ""],
    ["Target Market", "Parents of children ages 6-12", ""],
    ["Unique Value", "Instant, personalized, educational battle books", ""],
    ["", "", ""],
    ["PRICING TIERS", "Price", "Features"],
    ["Free", "$0", "8 animals (28 matchups) + 1 CYOA (Lion vs Tiger)"],
    ["Premium", "$9.99", "38 real animals + dinos + 28 CYOA matchups + PDF"],
    ["Ultimate", "$19.99", "All 47 animals + full CYOA (50/day) + PDF"],
    ["", "", ""],
    ["YEAR 1 PROJECTIONS", "Conservative", "Optimistic"],
    ["Total Users", "5,000", "20,000"],
    ["Free Tier", "4,250 (85%)", "16,000 (80%)"],
    ["Premium ($9.99)", "500 (10%)", "2,800 (14%)"],
    ["Ultimate ($19.99)", "250 (5%)", "1,200 (6%)"],
    ["Gross Revenue", "$9,990", "$51,800"],
    ["Total Costs", "$2,500", "$8,000"],
    ["Net Profit", "$7,490", "$43,800"],
    ["", "", ""],
    ["STARTUP COSTS", "Amount", "Notes"],
    ["Free Tier Pre-gen (28 books)", "$5", "One-time"],
    ["Free CYOA (Lion vs Tiger)", "$4", "27 paths"],
    ["Domain/Hosting", "$0", "Vercel free tier"],
    ["Total Startup", "$9", ""],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws1.merge_cells('A1:C1')
ws1.cell(1, 1).font = Font(bold=True, size=16)
style_header(ws1, 9, 3)
style_header(ws1, 14, 3)
style_header(ws1, 23, 3)
auto_width(ws1)

# ============================================================
# SHEET 2: Cost Structure
# ============================================================
ws2 = wb.create_sheet("Cost Structure")

cost_data = [
    ["COST STRUCTURE", "", "", ""],
    ["", "", "", ""],
    ["VARIABLE COSTS (Per Generation)", "Cost", "Notes", ""],
    ["FAL AI Image (per image)", "$0.003", "Flux Schnell", ""],
    ["OpenAI GPT-4o-mini (per book)", "$0.15", "Text generation", ""],
    ["", "", "", ""],
    ["STANDARD BOOK GENERATION", "Quantity", "Unit Cost", "Total"],
    ["Animal portraits (shared)", "8", "$0.003", "$0.024"],
    ["Battle images", "7", "$0.003", "$0.021"],
    ["OpenAI text", "1", "$0.15", "$0.15"],
    ["Total per NEW standard book", "", "", "$0.195"],
    ["", "", "", ""],
    ["CYOA GENERATION (Per Choice)", "Quantity", "Unit Cost", "Total"],
    ["Battle images", "2", "$0.003", "$0.006"],
    ["OpenAI text", "1", "$0.02", "$0.02"],
    ["Total per CYOA choice", "", "", "$0.026"],
    ["Full CYOA playthrough (5 choices)", "", "", "$0.13"],
    ["", "", "", ""],
    ["STORAGE COSTS (Vercel Blob)", "Size", "Monthly", ""],
    ["Per book (JSON + PDF)", "~4 MB", "", ""],
    ["1,000 cached books", "~4 GB", "$0.60", ""],
    ["10,000 cached books", "~40 GB", "$6.00", ""],
    ["Full CYOA cache (30K paths)", "~30 GB", "$4.50", ""],
    ["", "", "", ""],
    ["PRE-GENERATION COSTS", "Paths/Books", "Cost", "Notes"],
    ["Free tier books (28)", "28", "$5.46", "One-time"],
    ["Free CYOA (Lion vs Tiger)", "27", "$3.51", "One-time"],
    ["Premium CYOA (28 matchups)", "756", "$98.28", "Optional"],
    ["Total recommended pre-gen", "", "$9", "Free tier only"],
]

for row_idx, row_data in enumerate(cost_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws2.cell(1, 1).font = Font(bold=True, size=14)
style_header(ws2, 3, 4)
style_header(ws2, 7, 4)
style_header(ws2, 13, 4)
style_header(ws2, 19, 4)
style_header(ws2, 25, 4)
auto_width(ws2)

# ============================================================
# SHEET 3: Revenue Projections
# ============================================================
ws3 = wb.create_sheet("Revenue Projections")

# Monthly projections for Year 1
months = ["Month " + str(i) for i in range(1, 13)]
revenue_headers = ["Metric"] + months + ["Year 1 Total"]

revenue_data = [
    ["REVENUE PROJECTIONS - CONSERVATIVE SCENARIO", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    revenue_headers,
    ["New Free Users", 200, 250, 300, 350, 400, 450, 500, 550, 600, 500, 400, 350, 4850],
    ["New Premium Users", 15, 20, 30, 40, 50, 55, 60, 65, 70, 50, 35, 30, 520],
    ["New Ultimate Users", 5, 8, 15, 20, 25, 30, 35, 40, 40, 30, 20, 15, 283],
    ["Premium Revenue", 150, 200, 300, 400, 500, 550, 600, 650, 700, 500, 350, 300, 5200],
    ["Ultimate Revenue", 100, 160, 300, 400, 500, 600, 700, 800, 800, 600, 400, 300, 5660],
    ["Total Revenue", 250, 360, 600, 800, 1000, 1150, 1300, 1450, 1500, 1100, 750, 600, 10860],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["REVENUE PROJECTIONS - OPTIMISTIC SCENARIO", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    revenue_headers,
    ["New Free Users", 500, 750, 1000, 1500, 2000, 2500, 2500, 2000, 1500, 1000, 750, 500, 16500],
    ["New Premium Users", 50, 80, 150, 250, 400, 500, 500, 400, 300, 200, 100, 70, 3000],
    ["New Ultimate Users", 20, 40, 80, 150, 200, 250, 250, 200, 150, 100, 60, 40, 1540],
    ["Premium Revenue", 500, 800, 1500, 2500, 4000, 5000, 5000, 4000, 3000, 2000, 1000, 700, 30000],
    ["Ultimate Revenue", 400, 800, 1600, 3000, 4000, 5000, 5000, 4000, 3000, 2000, 1200, 800, 30800],
    ["Total Revenue", 900, 1600, 3100, 5500, 8000, 10000, 10000, 8000, 6000, 4000, 2200, 1500, 60800],
]

for row_idx, row_data in enumerate(revenue_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx in [7, 8, 9, 17, 18, 19] and col_idx > 1:
            cell.number_format = money_format

ws3.cell(1, 1).font = Font(bold=True, size=14)
ws3.cell(11, 1).font = Font(bold=True, size=14)
style_header(ws3, 3, 14)
style_header(ws3, 13, 14)
auto_width(ws3)

# ============================================================
# SHEET 4: Cost Projections
# ============================================================
ws4 = wb.create_sheet("Cost Projections")

cost_proj_data = [
    ["COST PROJECTIONS - CONSERVATIVE", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    revenue_headers,
    ["Standard Book Gens", 50, 60, 80, 100, 120, 130, 140, 150, 160, 130, 100, 80, 1300],
    ["Book Gen Cost ($0.20)", 10, 12, 16, 20, 24, 26, 28, 30, 32, 26, 20, 16, 260],
    ["CYOA Plays", 100, 150, 250, 400, 600, 800, 1000, 1200, 1300, 1000, 700, 500, 8000],
    ["CYOA Cache Hit %", "20%", "30%", "40%", "50%", "55%", "60%", "65%", "70%", "75%", "80%", "85%", "85%", ""],
    ["CYOA Gen Cost ($0.13)", 10, 14, 20, 26, 35, 42, 46, 47, 42, 26, 14, 10, 330],
    ["Storage (GB)", 1, 2, 3, 5, 7, 10, 13, 16, 20, 23, 25, 27, ""],
    ["Storage Cost", 0.15, 0.30, 0.45, 0.75, 1.05, 1.50, 1.95, 2.40, 3.00, 3.45, 3.75, 4.05, 23],
    ["Total Monthly Cost", 20, 26, 36, 47, 60, 70, 76, 79, 77, 55, 38, 30, 613],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["PROFIT ANALYSIS - CONSERVATIVE", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    revenue_headers,
    ["Revenue", 250, 360, 600, 800, 1000, 1150, 1300, 1450, 1500, 1100, 750, 600, 10860],
    ["Costs", 20, 26, 36, 47, 60, 70, 76, 79, 77, 55, 38, 30, 613],
    ["Net Profit", 230, 334, 564, 753, 940, 1080, 1224, 1371, 1423, 1045, 712, 570, 10247],
    ["Profit Margin", "92%", "93%", "94%", "94%", "94%", "94%", "94%", "95%", "95%", "95%", "95%", "95%", "94%"],
]

for row_idx, row_data in enumerate(cost_proj_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx in [5, 8, 10, 11, 16, 17, 18] and col_idx > 1 and col_idx < 14:
            if isinstance(value, (int, float)):
                cell.number_format = money_format

ws4.cell(1, 1).font = Font(bold=True, size=14)
ws4.cell(13, 1).font = Font(bold=True, size=14)
style_header(ws4, 3, 14)
style_header(ws4, 15, 14)
auto_width(ws4)

# ============================================================
# SHEET 5: Risk Analysis
# ============================================================
ws5 = wb.create_sheet("Risk Analysis")

risk_data = [
    ["RISK ANALYSIS & MITIGATIONS", "", "", ""],
    ["", "", "", ""],
    ["Risk", "Impact", "Likelihood", "Mitigation"],
    ["Low user acquisition", "High", "Medium", "SEO, content marketing, school partnerships"],
    ["Heavy CYOA abuser", "Medium", "Low", "50/day cap, fair usage policy, account bans"],
    ["Bot/script abuse", "High", "Low", "Rate limiting (5s), CAPTCHA, IP blocks"],
    ["FAL/OpenAI price increase", "Medium", "Low", "Cache aggressively, switch providers if needed"],
    ["Competitor enters market", "Medium", "Medium", "First-mover advantage, build brand loyalty"],
    ["Low conversion to paid", "High", "Medium", "Improve free tier demo, A/B test pricing"],
    ["", "", "", ""],
    ["MAXIMUM EXPOSURE ANALYSIS", "", "", ""],
    ["", "", "", ""],
    ["Scenario", "Calculation", "Max Cost", "Mitigation"],
    ["Single abuser (50/day, 30 days)", "1500 plays × $0.13 × 50% miss", "$97.50/month", "Account ban, worth it for cache"],
    ["10 heavy users", "10 × $97.50", "$975/month", "Covered by 50 Premium users"],
    ["Bot attack (1000 requests)", "Blocked by rate limiting", "$0", "Technical controls"],
    ["", "", "", ""],
    ["BREAK-EVEN ANALYSIS", "", "", ""],
    ["", "", "", ""],
    ["Metric", "Value", "", ""],
    ["Startup costs", "$9", "", ""],
    ["Monthly fixed costs", "~$5 (storage)", "", ""],
    ["Premium users needed to cover costs", "1", "", ""],
    ["Ultimate users needed for $1K/mo profit", "50", "", ""],
]

for row_idx, row_data in enumerate(risk_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws5.cell(1, 1).font = Font(bold=True, size=14)
ws5.cell(11, 1).font = Font(bold=True, size=14)
ws5.cell(18, 1).font = Font(bold=True, size=14)
style_header(ws5, 3, 4)
style_header(ws5, 13, 4)
style_header(ws5, 20, 4)
auto_width(ws5)

# ============================================================
# SHEET 6: Fair Usage Policy
# ============================================================
ws6 = wb.create_sheet("Fair Usage Policy")

policy_data = [
    ["FAIR USAGE POLICY - DRAFT", ""],
    ["", ""],
    ["Section", "Content"],
    ["1. Scope", "This Fair Usage Policy applies to all FightingBooks users with Premium or Ultimate subscriptions."],
    ["", ""],
    ["2. CYOA Usage Limits", "CYOA (Choose Your Own Adventure) mode is subject to reasonable usage limits:"],
    ["", "- Free tier: 1 matchup (Lion vs Tiger) only"],
    ["", "- Premium tier: 28 matchups (classic animals)"],
    ["", "- Ultimate tier: All matchups, up to 50 plays per day"],
    ["", ""],
    ["3. Prohibited Activities", "The following activities are prohibited:"],
    ["", "- Automated access, bots, or scripts"],
    ["", "- Systematic attempts to generate all possible content"],
    ["", "- Sharing account credentials"],
    ["", "- Circumventing technical limitations"],
    ["", ""],
    ["4. Enforcement", "FightingBooks reserves the right to:"],
    ["", "- Monitor usage patterns for abuse"],
    ["", "- Temporarily or permanently suspend accounts violating this policy"],
    ["", "- Modify usage limits with reasonable notice"],
    ["", "- Deny refunds for accounts terminated due to abuse"],
    ["", ""],
    ["5. Amendments", "This policy may be updated at any time. Continued use constitutes acceptance."],
    ["", ""],
    ["TECHNICAL CONTROLS", ""],
    ["", ""],
    ["Control", "Setting"],
    ["Daily CYOA cap", "50 plays/day"],
    ["Rate limiting", "5 seconds between plays"],
    ["Session limit", "CAPTCHA after 20 plays"],
    ["IP rate limit", "100 requests/hour"],
    ["Account required", "Email verification"],
]

for row_idx, row_data in enumerate(policy_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws6.cell(1, 1).font = Font(bold=True, size=14)
ws6.cell(26, 1).font = Font(bold=True, size=12)
style_header(ws6, 3, 2)
style_header(ws6, 28, 2)
ws6.column_dimensions['A'].width = 25
ws6.column_dimensions['B'].width = 80

# Save workbook
output_path = "/home/ubuntu/clawd/projects/fightingbooks-web/business-plan/FightingBooks_Financial_Model.xlsx"
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Saved to: {output_path}")
